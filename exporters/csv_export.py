import os
import pandas as pd
from datetime import datetime


class CSVExporter:
    """Export company leads to CSV and Excel with column ordering."""

    # Standard column order for all exports
    COLUMN_ORDER = [
        'name', 'sector', 'sub_sector', 'domain', 'email', 'phone',
        'country', 'city', 'region', 'employee_size',
        'source', 'source_url', 'outreach_status', 'notes'
    ]

    def __init__(self):
        os.makedirs("exports", exist_ok=True)

    def export_csv(self, records: list, sector: str = None, region: str = None) -> str | None:
        """Export records to a CSV file with proper column ordering."""
        if not records:
            print("No records to export.")
            return None

        df = pd.DataFrame(records)
        df = self._order_columns(df)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts = ["leads"]
        if sector:
            parts.append(sector)
        if region:
            parts.append(region)
        parts.append(timestamp)
        filename = f"exports/{'_'.join(parts)}.csv"

        df.to_csv(filename, index=False)
        print(f"Exported {len(records)} records to {filename}")
        return filename

    def export_excel(self, records: list, sector: str = None, region: str = None) -> str | None:
        """Export records to an Excel file with formatting."""
        if not records:
            print("No records to export.")
            return None

        df = pd.DataFrame(records)
        df = self._order_columns(df)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts = ["leads"]
        if sector:
            parts.append(sector)
        if region:
            parts.append(region)
        parts.append(timestamp)
        filename = f"exports/{'_'.join(parts)}.xlsx"

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Leads')
            
            worksheet = writer.sheets['Leads']
            from openpyxl.styles import PatternFill
            from openpyxl.utils import get_column_letter

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Determine column index of quality_score
            quality_col_idx = None
            for idx, col in enumerate(df.columns):
                if col == 'quality_score':
                    quality_col_idx = idx + 1
                    break

            if quality_col_idx is not None:
                # Iterate rows and apply color (skip header)
                for row_idx in range(2, len(df) + 2):
                    score_val = worksheet.cell(row=row_idx, column=quality_col_idx).value
                    if score_val is not None:
                        try:
                            score = int(score_val)
                            fill = green_fill if score >= 4 else (yellow_fill if score >= 2 else red_fill)
                            for col_idx in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = fill
                        except (ValueError, TypeError):
                            pass

            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                col_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[col_letter].width = min(max_len, 40)

        print(f"Exported {len(records)} records to {filename}")
        return filename

    def export_segmented(self, segments: dict) -> list:
        """Export each segment as a separate CSV file."""
        files = []
        for segment_name, records in segments.items():
            if not records:
                continue
            df = pd.DataFrame(records)
            df = self._order_columns(df)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"exports/{segment_name}_{timestamp}.csv"
            df.to_csv(filename, index=False)
            files.append(filename)
            print(f"  Segment '{segment_name}': {len(records)} records -> {filename}")
        return files

    def _order_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Reorder DataFrame columns to match the standard export order."""
        ordered = [c for c in self.COLUMN_ORDER if c in df.columns]
        extra = [c for c in df.columns if c not in self.COLUMN_ORDER]
        return df[ordered + extra]
